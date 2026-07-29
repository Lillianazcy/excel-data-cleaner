from __future__ import annotations

import re
import csv
import io
import sys
import unicodedata
from copy import copy
from pathlib import Path
from typing import Iterable
from zipfile import ZipFile

from openpyxl import Workbook, load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter


PREFECTURES = [
    "北海道", "青森県", "岩手県", "宮城県", "秋田県", "山形県", "福島県",
    "茨城県", "栃木県", "群馬県", "埼玉県", "千葉県", "東京都", "神奈川県",
    "新潟県", "富山県", "石川県", "福井県", "山梨県", "長野県", "岐阜県",
    "静岡県", "愛知県", "三重県", "滋賀県", "京都府", "大阪府", "兵庫県",
    "奈良県", "和歌山県", "鳥取県", "島根県", "岡山県", "広島県", "山口県",
    "徳島県", "香川県", "愛媛県", "高知県", "福岡県", "佐賀県", "長崎県",
    "熊本県", "大分県", "宮崎県", "鹿児島県", "沖縄県",
]

ADDRESS_MASTER = None


def app_dir() -> Path:
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent


def load_address_master():
    global ADDRESS_MASTER
    if ADDRESS_MASTER is not None:
        return ADDRESS_MASTER
    city_to_pref: dict[str, str] = {}
    town_to_pref: dict[str, set[str]] = {}
    candidates = [
        app_dir() / "ken_all.zip",
        Path.cwd() / "ken_all.zip",
        Path.home() / "Downloads" / "ken_all.zip",
    ]
    zip_path = next((p for p in candidates if p.exists()), None)
    if zip_path:
        try:
            with ZipFile(zip_path) as z:
                name = next(n for n in z.namelist() if n.lower().endswith(".csv"))
                text = io.TextIOWrapper(z.open(name), encoding="cp932", newline="")
                for row in csv.reader(text):
                    if len(row) < 9:
                        continue
                    pref = row[6].strip()
                    city = row[7].strip()
                    town = row[8].strip()
                    if pref and city:
                        city_to_pref.setdefault(city, pref)
                        for key in city_variants(city):
                            city_to_pref.setdefault(key, pref)
                    if pref and town and "以下に掲載がない場合" not in town:
                        town_to_pref.setdefault(town, set()).add(pref)
        except Exception:
            pass
    ADDRESS_MASTER = {"city_to_pref": city_to_pref, "town_to_pref": town_to_pref, "path": str(zip_path) if zip_path else ""}
    return ADDRESS_MASTER


def city_variants(city: str) -> set[str]:
    variants = {city}
    if city.startswith("東京都"):
        variants.add(city.replace("東京都", "", 1))
    for suffix in ("市", "郡", "村", "区"):
        pos = city.find(suffix)
        if pos >= 0:
            variants.add(city[:pos + len(suffix)])
    return {v for v in variants if v}


def lookup_prefecture_by_address(city_text, town_text="") -> str:
    master = load_address_master()
    city_map = master["city_to_pref"]
    town_map = master["town_to_pref"]
    texts = [clean_text(city_text), clean_text(town_text), to_half_width(city_text), to_half_width(town_text)]
    for text in texts:
        if not text:
            continue
        if text in city_map:
            return city_map[text]
        matches = sorted((k for k in city_map if k and k in text), key=len, reverse=True)
        if matches:
            return city_map[matches[0]]
    for text in texts:
        if not text:
            continue
        matches = sorted((k for k in town_map if k and k in text), key=len, reverse=True)
        if matches:
            prefs = town_map[matches[0]]
            if len(prefs) == 1:
                return next(iter(prefs))
    return ""


def load_sheet(path: str | Path):
    wb = load_workbook(path)
    return wb, wb.active


def save_workbook(wb, output_path: str | Path) -> None:
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    clean_workbook_metadata(wb)
    clone_workbook_for_export(wb).save(output)


def save_workbook_fast(wb, output_path: str | Path) -> None:
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    clean_workbook_metadata(wb)
    wb.save(output)


def clone_workbook_for_export(wb):
    clean = Workbook()
    clean.remove(clean.active)
    for ws in wb.worksheets:
        new_ws = clean.create_sheet(ws.title)
        for row in ws.iter_rows():
            for cell in row:
                new_cell = new_ws[cell.coordinate]
                new_cell.value = cell.value
                if cell.has_style:
                    new_cell._style = copy(cell._style)
                    new_cell.font = copy(cell.font)
                    new_cell.fill = copy(cell.fill)
                    new_cell.border = copy(cell.border)
                    new_cell.alignment = copy(cell.alignment)
                    new_cell.number_format = cell.number_format
                    new_cell.protection = copy(cell.protection)
        for key, dim in ws.column_dimensions.items():
            if dim.width:
                new_ws.column_dimensions[key].width = dim.width
        for key, dim in ws.row_dimensions.items():
            if dim.height:
                new_ws.row_dimensions[key].height = dim.height
        new_ws.sheet_view.showGridLines = ws.sheet_view.showGridLines
        new_ws.freeze_panes = ws.freeze_panes
        if ws.auto_filter.ref:
            new_ws.auto_filter.ref = ws.auto_filter.ref
        for merged in ws.merged_cells.ranges:
            new_ws.merge_cells(str(merged))
    try:
        clean.active = wb.index(wb.active)
    except Exception:
        pass
    return clean


def clean_workbook_metadata(wb) -> None:
    # Excel can show a repair prompt when stale workbook extension metadata remains
    # after heavy column insert/delete operations. The data sheet does not depend
    # on these records, so keep exports deliberately plain.
    try:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
    except Exception:
        pass
    try:
        wb._external_links = []
    except Exception:
        pass
    for ws in wb.worksheets:
        try:
            ws.sheet_properties.pageSetUpPr = None
        except Exception:
            pass


def idx(col: str) -> int:
    return column_index_from_string(col)


def value(ws, row: int, col: str):
    return ws.cell(row, idx(col)).value


def set_value(ws, row: int, col: str, val):
    cell = ws.cell(row, idx(col))
    cell.value = val
    return cell


def clean_text(v) -> str:
    if v is None:
        return ""
    return str(v)


def strip_quotes(v):
    if isinstance(v, str):
        return v.replace('"', "")
    return v


def to_half_width(v):
    if v is None:
        return ""
    return unicodedata.normalize("NFKC", str(v))


def text_format_columns(ws, columns: Iterable[str]) -> None:
    for col in columns:
        c = idx(col)
        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row, c)
            if cell.value is not None:
                cell.value = str(cell.value)
            cell.number_format = "@"


def replace_in_columns(ws, columns: Iterable[str], old: str, new: str) -> None:
    for col in columns:
        c = idx(col)
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row, c)
            if cell.value is not None:
                cell.value = str(cell.value).replace(old, new)


def replace_all_quotes(ws) -> None:
    for row in ws.iter_rows():
        for cell in row:
            cell.value = strip_quotes(cell.value)


def insert_columns(ws, before_col: str, amount: int, headers: list[str] | None = None) -> None:
    pos = idx(before_col)
    ws.insert_cols(pos, amount)
    if headers:
        for offset, header in enumerate(headers):
            ws.cell(1, pos + offset).value = header


def copy_column_visual_style(ws, source_col: str, target_col: str, fallback_col: str | None = None) -> None:
    source = idx(source_col)
    target = idx(target_col)
    fallback = idx(fallback_col) if fallback_col else None
    source_letter = get_column_letter(source)
    target_letter = get_column_letter(target)
    widths = []
    for col_letter in (source_letter, fallback_col):
        if col_letter:
            width = ws.column_dimensions[col_letter].width
            if width:
                widths.append(width)
    if widths:
        ws.column_dimensions[target_letter].width = max(widths)
    for row in range(1, ws.max_row + 1):
        src_cell = ws.cell(row, source)
        if fallback and not src_cell.has_style:
            src_cell = ws.cell(row, fallback)
        dst_cell = ws.cell(row, target)
        if src_cell.has_style:
            dst_cell._style = copy(src_cell._style)
            dst_cell.font = copy(src_cell.font)
            dst_cell.fill = copy(src_cell.fill)
            dst_cell.border = copy(src_cell.border)
            dst_cell.alignment = copy(src_cell.alignment)
            dst_cell.number_format = src_cell.number_format
            dst_cell.protection = copy(src_cell.protection)


def delete_columns(ws, columns: Iterable[str]) -> None:
    for col in sorted((idx(c) for c in columns), reverse=True):
        ws.delete_cols(col, 1)


def set_column_values(ws, col: str, values: list) -> None:
    c = idx(col)
    for row, val in enumerate(values, start=2):
        cell = ws.cell(row, c)
        cell.value = val
        cell.number_format = "@"


def column_values(ws, col: str) -> list:
    c = idx(col)
    return [ws.cell(row, c).value for row in range(2, ws.max_row + 1)]


def trim_single_spaces(text) -> str:
    return clean_text(text).strip()


def rotate_name(text) -> str:
    s = clean_text(text)
    if not s:
        return ""
    if " " not in s:
        return s
    first_space = s.find(" ")
    return (s + " " + s)[first_space + 1:first_space + 1 + len(s)]


def normalize_double_spaces(text) -> str:
    s = clean_text(text)
    while "  " in s:
        s = s.replace("  ", " ")
    return s


def flag_has_hyphen(text) -> str:
    return "@" if "-" in clean_text(text) else "X"


def add_hyphen_after_third_digit(text) -> str:
    s = to_half_width(text)
    if not s:
        return ""
    s = s.replace("－", "-")
    if "-" in s:
        return s
    digits = re.sub(r"\D", "", s)
    if len(digits) <= 3:
        return digits
    return digits[:3] + "-" + digits[3:]


def valid_prefecture(text) -> bool:
    s = clean_text(text)
    return any(s.endswith(suffix) for suffix in ("都", "道", "府", "県"))


def extract_prefecture(*texts) -> str:
    joined = " ".join(clean_text(t) for t in texts if t)
    for pref in PREFECTURES:
        if pref in joined:
            return pref
    return ""


def valid_city(text, pref: str = "") -> bool:
    s = clean_text(text)
    if not s:
        return False
    if s.endswith(("市", "郡", "村")):
        return True
    if pref == "東京都" and s.endswith("区"):
        return True
    return False


def extract_city(text, pref: str = "") -> str:
    s = clean_text(text)
    if not s:
        return ""
    if valid_city(s, pref):
        return s
    # Conservative extraction: keep text through the first municipality suffix.
    allowed = "市郡村区" if pref == "東京都" else "市郡村"
    pattern = rf"(.+?[{allowed}])"
    m = re.search(pattern, s)
    return m.group(1) if m else ""


def remove_test_rows(ws, columns: Iterable[str]) -> None:
    targets = [idx(c) for c in columns]
    needles = ("test", "テスト", "てすと")
    for row in range(ws.max_row, 1, -1):
        text = " ".join(clean_text(ws.cell(row, c).value).lower() for c in targets)
        if any(n in text for n in needles):
            ws.delete_rows(row, 1)


def remove_name_test_rows(ws) -> None:
    needles = ("テスト", "てすと", "test")
    for row in range(ws.max_row, 1, -1):
        text = clean_text(value(ws, row, "K")).lower()
        if any(n in text for n in needles):
            ws.delete_rows(row, 1)


def normalize_all_cells_to_half_width(ws) -> None:
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                cell.value = to_half_width(cell.value)


def to_half_width_katakana(text) -> str:
    s = to_half_width(text)
    reverse = {}
    for code in range(0xFF61, 0xFFA0):
        half = chr(code)
        full = unicodedata.normalize("NFKC", half)
        if full and full != half:
            reverse[full] = half
    # Handle common voiced/semi-voiced kana before single-character mapping.
    for code1 in range(0xFF61, 0xFFA0):
        for mark in ("ﾞ", "ﾟ"):
            half = chr(code1) + mark
            full = unicodedata.normalize("NFKC", half)
            if full and full != half:
                reverse[full] = half
    return "".join(reverse.get(ch, ch) for ch in s)


def ensure_prefecture(ws, pref_col: str, city_col: str, town_col: str, clear_if_city_empty: bool = False) -> None:
    for row in range(2, ws.max_row + 1):
        city = clean_text(value(ws, row, city_col))
        town = clean_text(value(ws, row, town_col))
        pref = clean_text(value(ws, row, pref_col))
        if clear_if_city_empty and not city:
            set_value(ws, row, pref_col, "")
            continue
        if not valid_prefecture(pref):
            found = extract_prefecture(pref, city, town)
            if not found:
                found = lookup_prefecture_by_address(city, town)
            if found:
                set_value(ws, row, pref_col, found)
        if not clean_text(value(ws, row, pref_col)):
            found = extract_prefecture(city, town) or lookup_prefecture_by_address(city, town)
            if found:
                set_value(ws, row, pref_col, found)


def split_city_tail_to_town(ws, city_col: str, town_col: str, pref_col: str) -> None:
    suffixes = ("市", "郡", "村", "区")
    city_map = load_address_master()["city_to_pref"]
    for row in range(2, ws.max_row + 1):
        city = clean_text(value(ws, row, city_col))
        pref = clean_text(value(ws, row, pref_col))
        if not city:
            continue
        exact_matches = sorted(
            (name for name, name_pref in city_map.items() if name and city.startswith(name) and (not pref or name_pref == pref)),
            key=len,
            reverse=True,
        )
        if exact_matches:
            head = exact_matches[0]
            tail = city[len(head):]
            if tail:
                set_value(ws, row, city_col, head)
                set_value(ws, row, town_col, tail + clean_text(value(ws, row, town_col)))
        else:
            allowed = suffixes if pref == "東京都" else ("市", "郡", "村")
            positions = [(city.rfind(s), s) for s in allowed if city.rfind(s) >= 0]
            if positions:
                pos, suffix = max(positions, key=lambda item: item[0])
                end = pos + len(suffix)
                if end < len(city):
                    head = city[:end]
                    tail = city[end:]
                    set_value(ws, row, city_col, head)
                    set_value(ws, row, town_col, tail + clean_text(value(ws, row, town_col)))
        city = clean_text(value(ws, row, city_col))
        city = re.sub(r"[A-Za-z0-9@._+\-]+", "", city)
        set_value(ws, row, city_col, city)


def clean_building_phone_text(text) -> str:
    s = clean_text(text)
    if not s:
        return ""
    s = re.sub(r"TEL\s*[:：]?\s*[\+\d][\d\-\s()（）]*", "", s, flags=re.IGNORECASE)
    s = re.sub(r"\+\s*\d[\d\-\s()（）]*", "", s)
    s = re.sub(r"(様方)\s*\d[\d\-\s()（）]*", r"\1", s)
    return s.strip()


def normalize_address_block(ws, pref_col: str, city_col: str, town_col: str, building_col: str, clear_pref_if_city_empty=False):
    for row in range(2, ws.max_row + 1):
        pref = clean_text(value(ws, row, pref_col))
        city = clean_text(value(ws, row, city_col))
        town = clean_text(value(ws, row, town_col))
        if clear_pref_if_city_empty and not city:
            set_value(ws, row, pref_col, "")
        elif not valid_prefecture(pref):
            found_pref = extract_prefecture(pref, city, town)
            if found_pref:
                set_value(ws, row, pref_col, found_pref)
                pref = found_pref
        if not valid_city(city, pref):
            found_city = extract_city(town, pref)
            if found_city:
                set_value(ws, row, city_col, found_city)
        for col in (pref_col, city_col, town_col, building_col):
            set_value(ws, row, col, to_half_width(value(ws, row, col)))


def normalize_phone_for_mobile(text) -> str:
    s = clean_text(text)
    if not s:
        return ""
    if s.startswith("'+81"):
        return "0" + s[4:]
    if s.startswith("+81"):
        return "0" + s[3:]
    if s.startswith("81"):
        return "0" + s[2:]
    return s


def format_phone_3_4_4(text) -> str:
    s = clean_text(text)
    if not s or "-" in s:
        return s
    digits = re.sub(r"\D", "", s)
    if len(digits) >= 11:
        return f"{digits[:3]}-{digits[3:7]}-{digits[7:11]}"
    return s


def normalize_tel(text) -> str:
    s = clean_text(text)
    if not s:
        return ""
    if "+0" in s:
        return s.replace("+", "")
    if "+81" in s:
        return s.replace("+81", "0")
    if s.startswith("81"):
        return "0" + s[2:]
    return s


def phone_digits(text) -> str:
    return re.sub(r"\D", "", clean_text(text))


def is_mobile_number(text) -> bool:
    digits = phone_digits(normalize_phone_for_mobile(text))
    return digits.startswith(("070", "080", "090"))


def same_without_hyphen(a, b) -> bool:
    return clean_text(a).replace("-", "") == clean_text(b).replace("-", "")


def stage1(input_path: str | Path, output_path: str | Path) -> str:
    wb, ws = load_sheet(input_path)
    replace_in_columns(ws, ["C", "E", "M"], ".", "/")
    ws.delete_cols(idx("BU"), 1)
    insert_columns(ws, "BC", 1, ["bank account type"])
    copy_column_visual_style(ws, "BB", "BC", "BD")
    ws.cell(1, idx("BC")).value = "bank account type"
    for row in range(2, ws.max_row + 1):
        set_value(ws, row, "CA", 1)
    replace_all_quotes(ws)
    insert_columns(ws, "BY", 2, ["temp_BY", "temp_BZ"])
    for row in range(2, ws.max_row + 1):
        bx = clean_text(value(ws, row, "BX"))
        ca = clean_text(value(ws, row, "CA"))
        set_value(ws, row, "BY", "" if bx == "" else f"1{bx}0")
        set_value(ws, row, "BZ", "" if ca == "" else f"{ca}1")
    needs_check = any(clean_text(value(ws, row, "BX")) and clean_text(value(ws, row, "BS")) for row in range(2, ws.max_row + 1))
    if needs_check:
        save_workbook(wb, output_path)
        return "stage1_check"
    _stage2_on_workbook(wb, ws)
    save_workbook(wb, output_path)
    return "stage2_check"


def stage2(input_path: str | Path, output_path: str | Path) -> None:
    wb, ws = load_sheet(input_path)
    _stage2_on_workbook(wb, ws)
    save_workbook(wb, output_path)


def _stage2_on_workbook(wb, ws) -> None:
    set_column_values(ws, "BX", column_values(ws, "BY"))
    set_column_values(ws, "CA", column_values(ws, "BZ"))
    delete_columns(ws, ["BY", "BZ"])
    for row in range(2, ws.max_row + 1):
        bs = clean_text(value(ws, row, "BS"))
        bx = clean_text(value(ws, row, "BX"))
        set_value(ws, row, "CI", 1 if bs else (1 if not bx else 2))
    insert_columns(ws, "L", 1, ["temp_L"])
    for row in range(2, ws.max_row + 1):
        set_value(ws, row, "L", trim_single_spaces(value(ws, row, "K")))
    set_column_values(ws, "K", column_values(ws, "L"))
    delete_columns(ws, ["L"])
    remove_name_test_rows(ws)
    normalize_all_cells_to_half_width(ws)
    insert_columns(ws, "L", 1, ["temp_L"])
    for row in range(2, ws.max_row + 1):
        set_value(ws, row, "L", rotate_name(value(ws, row, "K")))
    set_column_values(ws, "K", column_values(ws, "L"))
    delete_columns(ws, ["L"])


def stage3(input_path: str | Path, output_path: str | Path) -> None:
    wb, ws = load_sheet(input_path)
    for row in range(2, ws.max_row + 1):
        set_value(ws, row, "K", normalize_double_spaces(value(ws, row, "K")))
        set_value(ws, row, "K", to_half_width(value(ws, row, "K")))
    insert_columns(ws, "T", 1, ["temp_T"])
    copy_column_visual_style(ws, "S", "T", "U")
    ws.cell(1, idx("T")).value = "temp_T"
    for row in range(2, ws.max_row + 1):
        set_value(ws, row, "T", flag_has_hyphen(value(ws, row, "S")))
    save_workbook(wb, output_path)


def stage4(input_path: str | Path, output_path: str | Path) -> None:
    wb, ws = load_sheet(input_path)
    for row in range(2, ws.max_row + 1):
        if value(ws, row, "T") == "X":
            set_value(ws, row, "S", add_hyphen_after_third_digit(value(ws, row, "S")))
        set_value(ws, row, "S", to_half_width(value(ws, row, "S")))
    delete_columns(ws, ["T"])
    insert_columns(ws, "AH", 1, ["temp_AH"])
    for row in range(2, ws.max_row + 1):
        set_value(ws, row, "AH", flag_has_hyphen(value(ws, row, "AG")))
    for row in range(2, ws.max_row + 1):
        if value(ws, row, "AH") == "X":
            set_value(ws, row, "AG", add_hyphen_after_third_digit(value(ws, row, "AG")))
        set_value(ws, row, "AG", to_half_width(value(ws, row, "AG")))
    delete_columns(ws, ["AH"])
    remove_test_rows(ws, ["T", "U"])
    normalize_address_block(ws, "T", "U", "V", "W")
    ensure_prefecture(ws, "T", "U", "V")
    split_city_tail_to_town(ws, "U", "V", "T")
    ensure_prefecture(ws, "T", "U", "V")
    for row in range(2, ws.max_row + 1):
        for col in ("T", "U", "V"):
            set_value(ws, row, col, to_half_width(value(ws, row, col)))
        set_value(ws, row, "W", to_half_width_katakana(value(ws, row, "W")))

    normalize_address_block(ws, "AH", "AI", "AJ", "AK", clear_pref_if_city_empty=True)
    ensure_prefecture(ws, "AH", "AI", "AJ", clear_if_city_empty=True)
    split_city_tail_to_town(ws, "AI", "AJ", "AH")
    ensure_prefecture(ws, "AH", "AI", "AJ", clear_if_city_empty=True)
    for row in range(2, ws.max_row + 1):
        for col in ("AH", "AI", "AJ"):
            set_value(ws, row, col, to_half_width(value(ws, row, col)))
        set_value(ws, row, "AK", clean_building_phone_text(to_half_width_katakana(value(ws, row, "AK"))))
    save_workbook(wb, output_path)


def stage5(input_path: str | Path, output_path: str | Path) -> None:
    wb, ws = load_sheet(input_path)
    insert_columns(ws, "AL", 5, ["temp_AL", "temp_AM", "temp_AN", "temp_AO", "temp_AP"])
    for row in range(2, ws.max_row + 1):
        set_value(ws, row, "AL", value(ws, row, "AG") or value(ws, row, "S") or "")
        set_value(ws, row, "AM", value(ws, row, "AH") or value(ws, row, "T") or "")
        set_value(ws, row, "AN", value(ws, row, "AI") or value(ws, row, "U") or "")
        set_value(ws, row, "AO", value(ws, row, "AJ") or value(ws, row, "V") or "")
        set_value(ws, row, "AP", value(ws, row, "AK") if value(ws, row, "AJ") else (value(ws, row, "W") or ""))
    for src, dst in zip(["AL", "AM", "AN", "AO", "AP"], ["AG", "AH", "AI", "AJ", "AK"]):
        set_column_values(ws, dst, column_values(ws, src))
    delete_columns(ws, ["AL", "AM", "AN", "AO", "AP"])

    for row in range(2, ws.max_row + 1):
        ad = clean_text(value(ws, row, "AD"))
        if ad.startswith("81"):
            set_value(ws, row, "AD", "+81" + ad[2:])
        ae = normalize_phone_for_mobile(value(ws, row, "AD"))
        af = format_phone_3_4_4(ae)
        set_value(ws, row, "AD", af)

    for row in range(2, ws.max_row + 1):
        ab, ad = value(ws, row, "AB"), value(ws, row, "AD")
        if clean_text(ab) and is_mobile_number(ad) and clean_text(ab)[-4:] == clean_text(ad)[-4:] and clean_text(ab)[-9:] == clean_text(ad)[-9:]:
            set_value(ws, row, "AB", "")

    for row in range(2, ws.max_row + 1):
        ac = normalize_tel(value(ws, row, "AB"))
        ad = value(ws, row, "AD")
        set_value(ws, row, "AB", "" if not ac else ("" if is_mobile_number(ad) and (ac == ad or same_without_hyphen(ac, ad)) else ac))
        if clean_text(value(ws, row, "AD")) and not is_mobile_number(value(ws, row, "AD")):
            if not clean_text(value(ws, row, "AB")):
                set_value(ws, row, "AB", value(ws, row, "AD"))
            set_value(ws, row, "AD", "")

    for row in range(2, ws.max_row + 1):
        aq = normalize_tel(value(ws, row, "AP"))
        ar = value(ws, row, "AR")
        set_value(ws, row, "AP", "" if not aq else ("" if is_mobile_number(ar) and (aq == ar or same_without_hyphen(aq, ar)) else aq))
        if clean_text(value(ws, row, "AR")) and not is_mobile_number(value(ws, row, "AR")):
            if not clean_text(value(ws, row, "AP")):
                set_value(ws, row, "AP", value(ws, row, "AR"))
            set_value(ws, row, "AR", "")

    for row in range(2, ws.max_row + 1):
        ap, aq, ar = value(ws, row, "AP"), value(ws, row, "AQ"), value(ws, row, "AR")
        ab, ac, ad = value(ws, row, "AB"), value(ws, row, "AC"), value(ws, row, "AD")
        set_value(ws, row, "AP", ap or ab or "")
        set_value(ws, row, "AQ", aq or ac or "")
        set_value(ws, row, "AR", ar or ad or "")

    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                cell.value = to_half_width(cell.value)
                cell.number_format = "@"
    save_workbook(wb, output_path)


def find_sheet_case_insensitive(wb, name: str):
    wanted = name.lower()
    for ws in wb.worksheets:
        if ws.title.lower() == wanted:
            return ws
    return None


def main_customer_sheet(wb):
    ws = find_sheet_case_insensitive(wb, "CustomerData")
    if ws is not None:
        return ws
    for candidate in wb.worksheets:
        if candidate.title.lower() != "sheet1":
            return candidate
    return wb.active


def normalize_lookup_key(v):
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def is_zero_like(v) -> bool:
    if v in (None, ""):
        return False
    if isinstance(v, (int, float)):
        return v == 0
    return str(v).strip() == "0"


def build_sheet1_lookup(sheet1) -> dict[str, list]:
    lookup = {}
    rows_with_lookup_values = sorted(
        row
        for (row, col), cell in sheet1._cells.items()
        if col <= 6 and cell.value not in (None, "")
    )
    for row in rows_with_lookup_values:
        values = [sheet1.cell(row, col).value for col in range(1, 7)]
        key = normalize_lookup_key(values[0])
        if not key:
            continue
        lookup[key] = values
    return lookup


def delete_blank_rows(ws) -> None:
    cells_by_row = {}
    for (row, _col), cell in ws._cells.items():
        if row > 1:
            cells_by_row.setdefault(row, []).append(cell)
    for row in sorted(cells_by_row, reverse=True):
        cells = cells_by_row[row]
        if cells and all(cell.value in (None, "") for cell in cells):
            ws.delete_rows(row, 1)


def data_rows_by_id(ws) -> list[int]:
    a_col = idx("A")
    return sorted(
        row
        for (row, col), cell in ws._cells.items()
        if row >= 2 and col == a_col and cell.value not in (None, "")
    )


def remove_all_sheets_except(wb, keep_ws) -> None:
    for ws in list(wb.worksheets):
        if ws is not keep_ws:
            wb.remove(ws)
    wb.active = 0


def stage6(input_path: str | Path, output_path: str | Path) -> None:
    wb = load_workbook(input_path, data_only=False)
    ws = main_customer_sheet(wb)
    sheet1 = find_sheet_case_insensitive(wb, "Sheet1")
    if sheet1 is None:
        raise ValueError("Sheet1 was not found. Please add the manual Sheet1 page before running Stage 6.")
    lookup = build_sheet1_lookup(sheet1)
    for row in data_rows_by_id(ws):
        key = normalize_lookup_key(value(ws, row, "A"))
        found = lookup.get(key)
        if found:
            q_val = found[1]
            h_val = found[2]
            n_val = found[3]
            cb_val = found[4]
            cg_val = found[5]
        else:
            q_val, h_val, n_val, cb_val, cg_val = "", 0, "不明", 0, ""

        set_value(ws, row, "Q", "" if is_zero_like(q_val) else (q_val or ""))
        set_value(ws, row, "H", 0 if h_val in (None, "") else h_val)
        set_value(ws, row, "N", "不明" if n_val in (None, "") else n_val)
        set_value(ws, row, "CB", 0 if cb_val in (None, "") else cb_val)
        set_value(ws, row, "CG", "" if is_zero_like(cg_val) else (cg_val or ""))

    delete_blank_rows(ws)
    remove_all_sheets_except(wb, ws)
    save_workbook(wb, output_path)


STAGE_FUNCTIONS = {
    1: stage1,
    2: stage2,
    3: stage3,
    4: stage4,
    5: stage5,
    6: stage6,
}
